from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QTableWidget, QTableWidgetItem,
    QLineEdit, QPushButton, QLabel, QSizePolicy, QHeaderView, QFileDialog,
    QHBoxLayout, QMessageBox
)
from PyQt6.QtCore import Qt, QTimer
from PyQt6.QtGui import QIcon
import openpyxl

class UserTableWidget(QWidget):
    def __init__(self, auth_services):
        super().__init__()
        self.auth_services = auth_services
        self.init_ui()

    def init_ui(self):
        self.setStyleSheet("background-color: white;")

        # Layout principal
        layout = QVBoxLayout()
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(5)

        # Layout para filtro y botón de exportar
        filter_export_layout = QHBoxLayout()

        # Campo de filtro
        self.filter_input = QLineEdit(self)
        self.filter_input.setPlaceholderText("Filtrar por nombre de usuario o rol...")
        self.filter_input.textChanged.connect(self.filter_table)
        filter_export_layout.addWidget(self.filter_input)

        # Botón de exportar a Excel
        self.export_excel_btn = QPushButton(self)
        self.export_excel_btn.setIcon(QIcon.fromTheme("document-save-as"))
        self.export_excel_btn.setText("Exportar a Excel")
        self.export_excel_btn.clicked.connect(self.export_to_excel_dialog)
        filter_export_layout.addWidget(self.export_excel_btn)

        layout.addLayout(filter_export_layout)

        # Tabla
        self.table = QTableWidget(self)
        self.table.setColumnCount(3)
        self.table.setHorizontalHeaderLabels(["ID", "Nombre de Usuario", "Rol"])
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)

        # Configurar el encabezado para distribuir uniformemente
        header = self.table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)

        layout.addWidget(self.table, stretch=1)

        # Botones de operaciones
        self.refresh_btn = QPushButton("Refrescar Lista", self)
        self.refresh_btn.clicked.connect(self.load_users)
        layout.addWidget(self.refresh_btn)

        # Label para resultados
        self.result_label = QLabel(self)
        self.result_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(self.result_label)

        self.setLayout(layout)
        self.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)

        # Cargar usuarios inicialmente
        self.load_users()

        # Usar QTimer para retrasar la configuración del ancho de las columnas
        QTimer.singleShot(100, self.adjust_column_widths)

    def adjust_column_widths(self):
        table_width = self.table.viewport().width()
        column_count = self.table.columnCount()
        if column_count > 0 and table_width > 0:
            equal_width = int(table_width / column_count)
            for col_index in range(column_count):
                self.table.setColumnWidth(col_index, equal_width)
        else:
            print("Advertencia: No se pudo calcular el ancho de la tabla inicialmente, usando ancho por defecto para columnas.")
            default_column_width = 20
            for col_index in range(column_count):
                self.table.setColumnWidth(col_index, default_column_width)

    def load_users(self):
        """Cargar todos los usuarios en la tabla."""
        users = self.auth_services.get_all_users()
        self.table.setRowCount(0)

        for row, user in enumerate(users):
            self.table.insertRow(row)
            self.table.setItem(row, 0, QTableWidgetItem(str(user["id"])))
            self.table.setItem(row, 1, QTableWidgetItem(user["username"]))
            self.table.setItem(row, 2, QTableWidgetItem(user["role"]))

        self.adjust_column_widths()
        self.result_label.setText(f"Mostrando {len(users)} usuarios")

    def filter_table(self):
        """Filtrar la tabla basado en el texto del filtro."""
        filter_text = self.filter_input.text().lower()
        all_users = self.auth_services.get_all_users()

        filtered_users = [
            user for user in all_users
            if filter_text in user["username"].lower() or filter_text in user["role"].lower()
        ]

        self.table.setRowCount(0)
        for row, user in enumerate(filtered_users):
            self.table.insertRow(row)
            self.table.setItem(row, 0, QTableWidgetItem(str(user["id"])))
            self.table.setItem(row, 1, QTableWidgetItem(user["username"]))
            self.table.setItem(row, 2, QTableWidgetItem(user["role"]))

        self.result_label.setText(f"Mostrando {len(filtered_users)} usuarios filtrados")

    def export_to_excel_dialog(self):
        file_dialog = QFileDialog()
        file_path, _ = file_dialog.getSaveFileName(
            self,
            "Exportar a Excel",
            "",
            "Archivos Excel (*.xlsx)"
        )

        if file_path:
            if not file_path.lower().endswith(".xlsx"):
                file_path += ".xlsx"

            try:
                self.export_table_to_excel(file_path)
                QMessageBox.information(self, "Exportar a Excel", "Tabla exportada exitosamente a Excel.")
            except Exception as e:
                QMessageBox.critical(self, "Error al exportar", f"Ocurrió un error al exportar a Excel: {str(e)}")

    def export_table_to_excel(self, file_path):
        """Exporta los datos del QTableWidget a un archivo Excel (.xlsx)."""
        libro_excel = openpyxl.Workbook()
        hoja_excel = libro_excel.active

        # Encabezados
        for columna in range(self.table.columnCount()):
            encabezado = self.table.horizontalHeaderItem(columna).text()
            hoja_excel.cell(row=1, column=columna + 1, value=encabezado)

        # Datos de la tabla
        for fila in range(self.table.rowCount()):
            for columna in range(self.table.columnCount()):
                item = self.table.item(fila, columna)
                if item and item.text():
                    dato = item.text()
                    hoja_excel.cell(row=fila + 2, column=columna + 1, value=dato)

        libro_excel.save(file_path)
        print(f"Tabla exportada a Excel: {file_path}")
