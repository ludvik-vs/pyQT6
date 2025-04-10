from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit,
    QPushButton, QTableWidget, QTableWidgetItem, QDateEdit, QMessageBox,
    QFileDialog, QSizePolicy, QHeaderView
)
from PyQt6.QtCore import QDate, Qt, QTimer
from PyQt6.QtGui import QIcon
import openpyxl

class DiscountRangeForm(QWidget):
    def __init__(self, cashbox_service):
        super().__init__()
        self.cashbox_service = cashbox_service
        self.init_ui()

    def init_ui(self):
        # Layout principal
        main_layout = QVBoxLayout()

        # Layout para los campos de fecha y botones
        date_layout = QHBoxLayout()

        # Fecha de inicio
        self.start_date_label = QLabel("Fecha de Inicio:", self)
        self.start_date_input = QDateEdit(self)
        self.start_date_input.setCalendarPopup(True)
        self.start_date_input.setDate(QDate.currentDate())
        date_layout.addWidget(self.start_date_label)
        date_layout.addWidget(self.start_date_input)

        # Fecha de fin
        self.end_date_label = QLabel("Fecha de Fin:", self)
        self.end_date_input = QDateEdit(self)
        self.end_date_input.setCalendarPopup(True)
        self.end_date_input.setDate(QDate.currentDate())
        date_layout.addWidget(self.end_date_label)
        date_layout.addWidget(self.end_date_input)

        # Botón para buscar descuentos
        self.search_button = QPushButton("Buscar Descuentos", self)
        self.search_button.clicked.connect(self.search_discounts)
        date_layout.addWidget(self.search_button)

        # Botón para exportar a Excel
        self.export_excel_btn = QPushButton(self)
        self.export_excel_btn.setIcon(QIcon.fromTheme("document-save-as"))
        self.export_excel_btn.setText("Exportar a Excel")
        self.export_excel_btn.clicked.connect(self.export_to_excel_dialog)
        date_layout.addWidget(self.export_excel_btn)

        # Add refresh button
        self.refresh_btn = QPushButton("Refrescar Lista", self)
        self.refresh_btn.clicked.connect(self.refresh_discounts)
        date_layout.addWidget(self.refresh_btn)

        # Tabla para mostrar los resultados
        self.table = QTableWidget(self)
        self.table.setColumnCount(7)
        self.table.setHorizontalHeaderLabels([
            "ID", "Fecha", "User ID", "Order ID", "Monto Descuento",
            "Porcentaje Descuento", "Descripción"
        ])
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)

        header = self.table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)

        # Añadir elementos al layout principal
        main_layout.addLayout(date_layout)
        main_layout.addWidget(self.table)

        # Label para mostrar la suma de la columna 4
        self.total_discount_label = QLabel("Total Monto Descuento: 0.00", self)
        main_layout.addWidget(self.total_discount_label)

        self.setLayout(main_layout)

        # Initialize column widths
        QTimer.singleShot(100, self.adjust_column_widths)

    def adjust_column_widths(self):
        table_width = self.table.viewport().width()
        column_count = self.table.columnCount()
        if column_count > 0 and table_width > 0:
            equal_width = int(table_width / column_count)
            for col_index in range(column_count):
                self.table.setColumnWidth(col_index, equal_width)
        else:
            default_column_width = 20
            for col_index in range(column_count):
                self.table.setColumnWidth(col_index, default_column_width)

    def export_to_excel_dialog(self):
        file_dialog = QFileDialog()
        file_path, _ = file_dialog.getSaveFileName(
            self,
            "Exportar a Excel",
            "",
            "Archivos Excel (*.xlsx)"
        )

        if file_path:
            if not file_path.lower().endswith(".xlsx"):
                file_path += ".xlsx"

            try:
                self.export_table_to_excel(file_path)
                QMessageBox.information(self, "Exportar a Excel", "Tabla exportada exitosamente a Excel.")
            except Exception as e:
                QMessageBox.critical(self, "Error al exportar", f"Ocurrió un error al exportar a Excel: {str(e)}")

    def export_table_to_excel(self, file_path):
        """Exporta los datos del QTableWidget a un archivo Excel (.xlsx)."""
        libro_excel = openpyxl.Workbook()
        hoja_excel = libro_excel.active

        # Encabezados
        for columna in range(self.table.columnCount()):
            encabezado = self.table.horizontalHeaderItem(columna).text()
            hoja_excel.cell(row=1, column=columna + 1, value=encabezado)

        # Datos de la tabla
        for fila in range(self.table.rowCount()):
            for columna in range(self.table.columnCount()):
                item = self.table.item(fila, columna)
                if item and item.text():
                    dato = item.text()
                    hoja_excel.cell(row=fila + 2, column=columna + 1, value=dato)

        libro_excel.save(file_path)

    def search_discounts(self):
        """Buscar descuentos en el rango de fechas seleccionado."""
        start_date = self.start_date_input.date().toString("yyyy-MM-dd")
        end_date = self.end_date_input.date().toString("yyyy-MM-dd")

        try:
            discounts = self.cashbox_service.get_discounts_in_date_range(start_date, end_date)
            self.display_discounts(discounts)
            self.update_total_discount_label(discounts)
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al obtener descuentos: {e}")

    def display_discounts(self, discounts):
        """Mostrar los descuentos en la tabla."""
        self.table.setRowCount(len(discounts))

        for row, discount in enumerate(discounts):
            self.table.setItem(row, 0, QTableWidgetItem(str(discount[0])))
            self.table.setItem(row, 1, QTableWidgetItem(discount[1]))
            self.table.setItem(row, 2, QTableWidgetItem(str(discount[2])))
            self.table.setItem(row, 3, QTableWidgetItem(str(discount[3])))
            self.table.setItem(row, 4, QTableWidgetItem(str(discount[4])))
            self.table.setItem(row, 5, QTableWidgetItem(f"{discount[5]:.2f}%"))
            self.table.setItem(row, 6, QTableWidgetItem(discount[6]))

    def refresh_discounts(self):
        """Refresh the discounts list using current date range."""
        self.search_discounts()
        self.adjust_column_widths()

    def update_total_discount_label(self, discounts):
        """Actualiza el label con la suma de la columna 'Monto Descuento'."""
        total_discount = sum(float(discount[4]) for discount in discounts)
        self.total_discount_label.setText(f"Total Monto Descuento: {total_discount:.2f}")
