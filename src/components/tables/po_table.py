from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QTableWidget, QTableWidgetItem,
    QLineEdit, QPushButton, QLabel, QSizePolicy, QHeaderView,
    QFileDialog, QHBoxLayout, QMessageBox
)
from PyQt6.QtCore import Qt, QTimer
from PyQt6.QtGui import QIcon
import openpyxl

class ProductionOrderTable(QWidget):
    def __init__(self, production_order_service):
        super().__init__()
        self.production_order_service = production_order_service
        self.init_ui()

    def init_ui(self):
        self.setStyleSheet("background-color: white;")

        layout = QVBoxLayout()
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(5)

        filter_export_layout = QHBoxLayout()

        self.filter_input = QLineEdit(self)
        self.filter_input.setPlaceholderText("Filtrar por ID de Orden o Cliente...")
        self.filter_input.textChanged.connect(self.filter_table)
        filter_export_layout.addWidget(self.filter_input)

        self.export_excel_btn = QPushButton(self)
        self.export_excel_btn.setIcon(QIcon.fromTheme("document-save-as"))
        self.export_excel_btn.setText("Exportar a Excel")
        self.export_excel_btn.clicked.connect(self.export_to_excel_dialog)
        filter_export_layout.addWidget(self.export_excel_btn)

        layout.addLayout(filter_export_layout)

        self.table = QTableWidget(self)
        self.table.setColumnCount(11)
        self.table.setHorizontalHeaderLabels([
            "ID", "ID Orden de Trabajo", "Fecha Inicio", "Fecha Fin", "ID Colaborador",
            "ID Cliente", "ID Producto", "Cantidad", "Estado", "Detalles de Tareas", "Nota"
        ])
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)

        header = self.table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)

        layout.addWidget(self.table, stretch=1)

        self.refresh_btn = QPushButton("Refrescar Lista", self)
        self.refresh_btn.clicked.connect(self.load_production_orders)
        layout.addWidget(self.refresh_btn)

        self.result_label = QLabel(self)
        self.result_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(self.result_label)

        self.setLayout(layout)
        self.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)

        QTimer.singleShot(100, self.adjust_column_widths)
        self.load_production_orders()

    def adjust_column_widths(self):
        table_width = self.table.viewport().width()
        column_count = self.table.columnCount()
        if column_count > 0 and table_width > 0:
            equal_width = int(table_width / column_count)
            for col_index in range(column_count):
                self.table.setColumnWidth(col_index, equal_width)
        else:
            default_column_width = 20
            for col_index in range(column_count):
                self.table.setColumnWidth(col_index, default_column_width)

    def load_production_orders(self):
        production_orders = self.production_order_service.get_all_orders()
        self.table.setRowCount(0)

        for row, order in enumerate(production_orders):
            self.table.insertRow(row)
            self.table.setItem(row, 0, QTableWidgetItem(str(order["id"])))
            self.table.setItem(row, 1, QTableWidgetItem(str(order["work_order_id"])))
            self.table.setItem(row, 2, QTableWidgetItem(order["start_date"]))
            self.table.setItem(row, 3, QTableWidgetItem(order["end_date"]))
            self.table.setItem(row, 4, QTableWidgetItem(str(order["colaborador_id"])))
            self.table.setItem(row, 5, QTableWidgetItem(str(order["client_id"])))
            self.table.setItem(row, 6, QTableWidgetItem(str(order["product_id"])))
            self.table.setItem(row, 7, QTableWidgetItem(str(order["quantity"])))
            self.table.setItem(row, 8, QTableWidgetItem(order["order_status"]))
            self.table.setItem(row, 9, QTableWidgetItem(order["tasks_details"]))
            self.table.setItem(row, 10, QTableWidgetItem(order["note"]))

        self.adjust_column_widths()
        self.result_label.setText(f"Mostrando {len(production_orders)} órdenes de producción")

    def filter_table(self):
        filter_text = self.filter_input.text().lower()
        all_production_orders = self.production_order_service.get_all_orders()

        filtered_production_orders = [
            order for order in all_production_orders
            if filter_text in str(order["work_order_id"]).lower() or filter_text in str(order["client_id"]).lower()
        ]

        self.table.setRowCount(0)

        for row, order in enumerate(filtered_production_orders):
            self.table.insertRow(row)
            self.table.setItem(row, 0, QTableWidgetItem(str(order["id"])))
            self.table.setItem(row, 1, QTableWidgetItem(str(order["work_order_id"])))
            self.table.setItem(row, 2, QTableWidgetItem(order["start_date"]))
            self.table.setItem(row, 3, QTableWidgetItem(order["end_date"]))
            self.table.setItem(row, 4, QTableWidgetItem(str(order["colaborador_id"])))
            self.table.setItem(row, 5, QTableWidgetItem(str(order["client_id"])))
            self.table.setItem(row, 6, QTableWidgetItem(str(order["product_id"])))
            self.table.setItem(row, 7, QTableWidgetItem(str(order["quantity"])))
            self.table.setItem(row, 8, QTableWidgetItem(order["order_status"]))
            self.table.setItem(row, 9, QTableWidgetItem(order["tasks_details"]))
            self.table.setItem(row, 10, QTableWidgetItem(order["note"]))

        self.result_label.setText(f"Mostrando {len(filtered_production_orders)} órdenes de producción filtradas")

    def export_to_excel_dialog(self):
        file_dialog = QFileDialog()
        file_path, _ = file_dialog.getSaveFileName(
            self,
            "Exportar a Excel",
            "",
            "Archivos Excel (*.xlsx)"
        )

        if file_path:
            if not file_path.lower().endswith(".xlsx"):
                file_path += ".xlsx"

            try:
                self.export_table_to_excel(file_path)
                QMessageBox.information(self, "Exportar a Excel", "Tabla exportada exitosamente a Excel.")
            except Exception as e:
                QMessageBox.critical(self, "Error al exportar", f"Ocurrió un error al exportar a Excel: {str(e)}")

    def export_table_to_excel(self, file_path):
        """Exporta los datos del QTableWidget a un archivo Excel (.xlsx)."""
        libro_excel = openpyxl.Workbook()
        hoja_excel = libro_excel.active

        # Encabezados
        for columna in range(self.table.columnCount()):
            encabezado = self.table.horizontalHeaderItem(columna).text()
            hoja_excel.cell(row=1, column=columna + 1, value=encabezado)

        # Datos de la tabla
        for fila in range(self.table.rowCount()):
            for columna in range(self.table.columnCount()):
                item = self.table.item(fila, columna)
                if item and item.text():
                    dato = item.text()
                    hoja_excel.cell(row=fila + 2, column=columna + 1, value=dato)

        libro_excel.save(file_path)
        print(f"Tabla exportada a Excel: {file_path}")
