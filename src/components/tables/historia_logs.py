from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QTableWidget, QTableWidgetItem,
    QPushButton, QLabel, QSizePolicy, QHeaderView,
    QFileDialog, QHBoxLayout, QMessageBox, QDateTimeEdit
)
from PyQt6.QtCore import Qt, QTimer, QDateTime
from PyQt6.QtGui import QIcon
from src.services.logs_services import LogsServices
import openpyxl

class LogsHistoryTable(QWidget):
    def __init__(self, logs_service: LogsServices):
        super().__init__()
        self.logs_service = logs_service
        self.init_ui()

    def init_ui(self):
        self.setStyleSheet("background-color: white;")
        layout = QVBoxLayout()
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(5)

        # Date range filter layout
        date_filter_layout = QHBoxLayout()

        # Start date filter
        start_date_label = QLabel("Fecha Inicio:", self)
        self.start_date = QDateTimeEdit(self)
        self.start_date.setDateTime(QDateTime.currentDateTime().addDays(-7))
        self.start_date.setCalendarPopup(True)
        date_filter_layout.addWidget(start_date_label)
        date_filter_layout.addWidget(self.start_date)

        # End date filter
        end_date_label = QLabel("Fecha Fin:", self)
        self.end_date = QDateTimeEdit(self)
        self.end_date.setDateTime(QDateTime.currentDateTime())
        self.end_date.setCalendarPopup(True)
        date_filter_layout.addWidget(end_date_label)
        date_filter_layout.addWidget(self.end_date)

        # Filter button
        self.filter_btn = QPushButton("Filtrar", self)
        self.filter_btn.clicked.connect(self.filter_logs)
        date_filter_layout.addWidget(self.filter_btn)

        # Export button
        self.export_excel_btn = QPushButton("Exportar a Excel", self)
        self.export_excel_btn.clicked.connect(self.export_to_excel_dialog)
        date_filter_layout.addWidget(self.export_excel_btn)

        layout.addLayout(date_filter_layout)

        # Logs table
        self.table = QTableWidget(self)
        self.table.setColumnCount(4)
        self.table.setHorizontalHeaderLabels(["ID", "Usuario", "Actividad", "Fecha y Hora"])
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)

        # Configure header
        header = self.table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)

        layout.addWidget(self.table, stretch=1)

        # Results label
        self.result_label = QLabel(self)
        self.result_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(self.result_label)

        self.setLayout(layout)
        self.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)

        # Initial load
        self.filter_logs()
        QTimer.singleShot(100, self.adjust_column_widths)

    def adjust_column_widths(self):
        table_width = self.table.viewport().width()
        column_count = self.table.columnCount()
        if column_count > 0 and table_width > 0:
            equal_width = int(table_width / column_count)
            for col_index in range(column_count):
                self.table.setColumnWidth(col_index, equal_width)
        else:
            default_column_width = 20
            for col_index in range(column_count):
                self.table.setColumnWidth(col_index, default_column_width)

        # Configure header to stretch
        header = self.table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        header.setStretchLastSection(True)

    def filter_logs(self):
        """Filter logs based on selected date range"""
        start_date = self.start_date.dateTime().toString("yyyy-MM-dd HH:mm:ss")
        end_date = self.end_date.dateTime().toString("yyyy-MM-dd HH:mm:ss")
        
        logs_json = self.logs_service.get_logs_by_date_range_json(start_date, end_date)
        import json
        logs = json.loads(logs_json)
        self.table.setRowCount(0)

        for row, log in enumerate(logs):
            self.table.insertRow(row)
            self.table.setItem(row, 0, QTableWidgetItem(str(log['id'])))  # ID
            self.table.setItem(row, 1, QTableWidgetItem(str(log['username'])))  # Username
            self.table.setItem(row, 2, QTableWidgetItem(str(log['activity'])))  # Activity
            self.table.setItem(row, 3, QTableWidgetItem(str(log['timestamp'])))  # DateTime

        self.adjust_column_widths()
        self.result_label.setText(f"Mostrando {len(logs)} registros")

    def export_to_excel_dialog(self):
        """Open dialog to export logs to Excel"""
        file_dialog = QFileDialog()
        file_path, _ = file_dialog.getSaveFileName(
            self,
            "Exportar a Excel",
            "",
            "Archivos Excel (*.xlsx)"
        )

        if file_path:
            if not file_path.lower().endswith(".xlsx"):
                file_path += ".xlsx"

            try:
                self.export_table_to_excel(file_path)
                QMessageBox.information(self, "Exportar a Excel", "Registros exportados exitosamente a Excel.")
            except Exception as e:
                QMessageBox.critical(self, "Error al exportar", f"Ocurrió un error al exportar a Excel: {str(e)}")

    def export_table_to_excel(self, file_path):
        """Export the table data to Excel file"""
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # Headers
        headers = ["ID", "Usuario", "Actividad", "Fecha y Hora"]
        for col, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col, value=header)

        # Data
        for row in range(self.table.rowCount()):
            for col in range(self.table.columnCount()):
                item = self.table.item(row, col)
                if item and item.text():
                    sheet.cell(row=row + 2, column=col + 1, value=item.text())

        workbook.save(file_path)