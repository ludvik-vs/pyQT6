from PyQt6.QtWidgets import (
    QWidget, QTableWidget, QTableWidgetItem, QVBoxLayout, QHeaderView, QDialog, 
    QLabel, QPushButton, QHBoxLayout, QTableWidgetItem, QMessageBox, QFileDialog
)
from PyQt6.QtCore import Qt
import pandas as pd
# Add openpyxl import
import openpyxl

class CashBoxResumeTableWidget(QWidget):
    def __init__(
        self,
        current_user_data,
        aunth_service,
        client_service,
        colaborator_service,
        work_order_service,
        production_order_service,
        cashbox_service
    ):
        super().__init__()
        self.current_user_data = current_user_data
        self.aunth_service = aunth_service
        self.client_service = client_service
        self.colaborator_service = colaborator_service
        self.work_order_service = work_order_service
        self.production_order_service = production_order_service
        self.cashbox_service = cashbox_service

        self.init_ui()

    def init_ui(self):
        self.layout = QVBoxLayout()

        # Create the table widget
        self.table_widget = QTableWidget()
        self.table_widget.setColumnCount(2)
        self.table_widget.setHorizontalHeaderLabels(["Fecha", "Índice"])
        self.table_widget.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.table_widget.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table_widget.cellDoubleClicked.connect(self.show_detail_popup)

        self.layout.addWidget(self.table_widget)
        self.setLayout(self.layout)

        # Load data into the table
        self.load_data()

    def load_data(self):
        # Fetch data from the cashbox service
        data = self.cashbox_service.get_all_index_identifiers_service()

        self.table_widget.setRowCount(len(data))
        for row_idx, record in enumerate(data):
            self.table_widget.setItem(row_idx, 0, QTableWidgetItem(record[1]))
            self.table_widget.setItem(row_idx, 1, QTableWidgetItem(str(record[2])))

    # Move show_detail_popup inside the class
    def show_detail_popup(self, row, column):
        record = {
            "fecha": self.table_widget.item(row, 0).text(),
            "index_identifier": self.table_widget.item(row, 1).text(),
        }
        
        count_data = self.cashbox_service.get_cash_count_denomination_by_index_identifier_service(
            record["index_identifier"])
        
        detail_popup = CashCountDetailPopup(record, count_data)
        detail_popup.exec()

class CashCountDetailPopup(QDialog):
    def __init__(self, record, count_data=None):
        super().__init__()
        self.record = record
        self.count_data = count_data
        self.init_ui()

    def init_ui(self):
        self.setWindowTitle("Detalle de Arqueo de Efectivo")
        self.setMinimumWidth(800)
        self.layout = QVBoxLayout()

        # Display record details
        header_layout = QHBoxLayout()
        header_layout.addWidget(QLabel(f"Fecha: {self.record['fecha']}"))
        header_layout.addWidget(QLabel(f"Índice: {self.record['index_identifier']}"))
        self.layout.addLayout(header_layout)

        # Create table
        self.table = QTableWidget()
        self.table.setColumnCount(9)
        headers = ["ID", "Usuario", "Índice", "Fecha", "Denominación NIO", 
                  "Denominación USD", "Tipo Cambio", "Cantidad", "Subtotal"]
        self.table.setHorizontalHeaderLabels(headers)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)

        # Populate table
        if self.count_data:
            self.table.setRowCount(len(self.count_data))
            for row, data in enumerate(self.count_data):
                for col, value in enumerate(data):
                    item = QTableWidgetItem(str(value) if value is not None else "")
                    self.table.setItem(row, col, item)

        self.layout.addWidget(self.table)

        # Buttons layout
        button_layout = QHBoxLayout()
        
        # Export button
        export_button = QPushButton("Exportar a Excel")
        export_button.clicked.connect(self.export_to_excel)
        button_layout.addWidget(export_button)

        # Close button
        close_button = QPushButton("Cerrar")
        close_button.clicked.connect(self.close)
        button_layout.addWidget(close_button)

        self.layout.addLayout(button_layout)
        self.setLayout(self.layout)

    def export_to_excel(self):
        """Export cash count details to Excel"""
        file_dialog = QFileDialog()
        suggested_name = f"Arqueo_{self.record['index_identifier']}_{self.record['fecha']}.xlsx"
        file_path, _ = file_dialog.getSaveFileName(
            self,
            "Exportar a Excel",
            suggested_name,
            "Archivos Excel (*.xlsx)"
        )

        if file_path:
            if not file_path.lower().endswith(".xlsx"):
                file_path += ".xlsx"

            try:
                self.export_table_to_excel(file_path)
                QMessageBox.information(self, "Exportar a Excel", "Arqueo exportado exitosamente a Excel.")
            except Exception as e:
                QMessageBox.critical(self, "Error al exportar", f"Ocurrió un error al exportar a Excel: {str(e)}")

    def export_table_to_excel(self, file_path):
        """Exports QTableWidget data to Excel file (.xlsx)"""
        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        # Add header with cash count info
        worksheet.cell(row=1, column=1, value=f"Fecha: {self.record['fecha']}")
        worksheet.cell(row=1, column=3, value=f"Índice: {self.record['index_identifier']}")

        # Table headers (row 3)
        for column in range(self.table.columnCount()):
            header = self.table.horizontalHeaderItem(column).text()
            worksheet.cell(row=3, column=column + 1, value=header)

        # Table data (starting from row 4)
        for row in range(self.table.rowCount()):
            for column in range(self.table.columnCount()):
                item = self.table.item(row, column)
                if item and item.text():
                    value = item.text()
                    worksheet.cell(row=row + 4, column=column + 1, value=value)

        workbook.save(file_path)
