from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QTableWidget, QTableWidgetItem,
    QLineEdit, QPushButton, QLabel, QSizePolicy, QHeaderView,
    QFileDialog, QHBoxLayout, QMessageBox, QApplication
)
from PyQt6.QtCore import Qt, QTimer
from PyQt6.QtGui import QIcon
from src.services.rh_service import ColaboratorService
import openpyxl

class ColaboratorTableWidget(QWidget):
    def __init__(self, colaborator_service: ColaboratorService):
        super().__init__()
        self.colaborator_service = colaborator_service
        self.init_ui()

    def init_ui(self):
        self.setStyleSheet("background-color: white;")

        layout = QVBoxLayout()
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(5)

        filter_export_layout = QHBoxLayout()

        self.filter_input = QLineEdit(self)
        self.filter_input.setPlaceholderText("Filtrar por nombre o apellido...")
        self.filter_input.textChanged.connect(self.filter_table)
        filter_export_layout.addWidget(self.filter_input)

        self.export_excel_btn = QPushButton(self)
        self.export_excel_btn.setIcon(QIcon.fromTheme("document-save-as"))
        self.export_excel_btn.setText("Exportar a Excel")
        self.export_excel_btn.clicked.connect(self.export_to_excel_dialog)
        filter_export_layout.addWidget(self.export_excel_btn)

        layout.addLayout(filter_export_layout)

        self.table = QTableWidget(self)
        self.table.setColumnCount(15)
        self.table.setHorizontalHeaderLabels([
            "ID", "Nombre", "Apellido", "Teléfono Personal", "Documento Identidad",
            "Fecha Ingreso", "Contacto Emergencia", "Teléfono Emergencia", "Fecha de Baja",
            "Salario", "Activo", "Puesto", "Fecha Nacimiento", "Número Seguro Social", "Información Adicional"
        ])
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)

        header = self.table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)

        layout.addWidget(self.table, stretch=1)

        self.refresh_btn = QPushButton("Refrescar Lista", self)
        self.refresh_btn.clicked.connect(self.load_colaborators)
        layout.addWidget(self.refresh_btn)

        self.result_label = QLabel(self)
        self.result_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(self.result_label)

        self.setLayout(layout)
        self.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)

        QTimer.singleShot(100, self.adjust_column_widths)
        self.load_colaborators()

    def adjust_column_widths(self):
        table_width = self.table.viewport().width()
        column_count = self.table.columnCount()
        if column_count > 0 and table_width > 0:
            equal_width = int(table_width / column_count)
            for col_index in range(column_count):
                self.table.setColumnWidth(col_index, equal_width)
        else:
            default_column_width = 20
            for col_index in range(column_count):
                self.table.setColumnWidth(col_index, default_column_width)

    def load_colaborators(self):
        colaborators = self.colaborator_service.get_all_colaborators()
        self.table.setRowCount(0)

        if colaborators and isinstance(colaborators[0], tuple):
            colaborator_keys = [
                "id", "nombre", "apellido", "telefono_personal", "documento_identidad",
                "fecha_ingreso", "nombre_contacto_emergencia", "telefono_emergencia", "fecha_baja",
                "salario", "is_active", "puesto", "fecha_nacimiento", "numero_seguro_social", "informacion_adicional"
            ]
            colaborators = [dict(zip(colaborator_keys, colaborator)) for colaborator in colaborators]

        for row, colaborator in enumerate(colaborators):
            self.table.insertRow(row)
            self.table.setItem(row, 0, QTableWidgetItem(str(colaborator["id"])))
            self.table.setItem(row, 1, QTableWidgetItem(colaborator["nombre"]))
            self.table.setItem(row, 2, QTableWidgetItem(colaborator["apellido"]))
            self.table.setItem(row, 3, QTableWidgetItem(colaborator["telefono_personal"] or ""))
            self.table.setItem(row, 4, QTableWidgetItem(colaborator["documento_identidad"] or ""))
            self.table.setItem(row, 5, QTableWidgetItem(colaborator["fecha_ingreso"]))
            self.table.setItem(row, 6, QTableWidgetItem(colaborator["nombre_contacto_emergencia"] or ""))
            self.table.setItem(row, 7, QTableWidgetItem(colaborator["telefono_emergencia"] or ""))
            self.table.setItem(row, 8, QTableWidgetItem(colaborator["fecha_baja"]))
            self.table.setItem(row, 9, QTableWidgetItem(str(colaborator["salario"])))
            self.table.setItem(row, 10, QTableWidgetItem(str(int(colaborator["is_active"]))))
            self.table.setItem(row, 11, QTableWidgetItem(colaborator["puesto"] or ""))
            self.table.setItem(row, 12, QTableWidgetItem(colaborator["fecha_nacimiento"] or ""))
            self.table.setItem(row, 13, QTableWidgetItem(colaborator["numero_seguro_social"] or ""))
            self.table.setItem(row, 14, QTableWidgetItem(colaborator["informacion_adicional"]))

        self.adjust_column_widths()
        self.result_label.setText(f"Mostrando {len(colaborators)} colaboradores")

    def filter_table(self):
        filter_text = self.filter_input.text().lower()
        all_colaborators = self.colaborator_service.get_all_colaborators()

        if all_colaborators and isinstance(all_colaborators[0], tuple):
            colaborator_keys = [
                "id", "nombre", "apellido", "telefono_personal", "documento_identidad",
                "fecha_ingreso", "nombre_contacto_emergencia", "telefono_emergencia", "fecha_baja",
                "salario", "is_active", "puesto", "fecha_nacimiento", "numero_seguro_social", "informacion_adicional"
            ]
            all_colaborators = [dict(zip(colaborator_keys, colaborator)) for colaborator in all_colaborators]

        filtered_colaborators = [
            colaborator for colaborator in all_colaborators
            if filter_text in colaborator["nombre"].lower() or filter_text in colaborator["apellido"].lower()
        ]

        self.table.setRowCount(0)

        for row, colaborator in enumerate(filtered_colaborators):
            self.table.insertRow(row)
            self.table.setItem(row, 0, QTableWidgetItem(str(colaborator["id"])))
            self.table.setItem(row, 1, QTableWidgetItem(colaborator["nombre"]))
            self.table.setItem(row, 2, QTableWidgetItem(colaborator["apellido"]))
            self.table.setItem(row, 3, QTableWidgetItem(colaborator["telefono_personal"] or ""))
            self.table.setItem(row, 4, QTableWidgetItem(colaborator["documento_identidad"] or ""))
            self.table.setItem(row, 5, QTableWidgetItem(colaborator["fecha_ingreso"]))
            self.table.setItem(row, 6, QTableWidgetItem(colaborator["nombre_contacto_emergencia"] or ""))
            self.table.setItem(row, 7, QTableWidgetItem(colaborator["telefono_emergencia"] or ""))
            self.table.setItem(row, 8, QTableWidgetItem(colaborator["fecha_baja"]))
            self.table.setItem(row, 9, QTableWidgetItem(str(colaborator["salario"])))
            self.table.setItem(row, 10, QTableWidgetItem(colaborator["is_active"]))
            self.table.setItem(row, 11, QTableWidgetItem(colaborator["puesto"] or ""))
            self.table.setItem(row, 12, QTableWidgetItem(colaborator["fecha_nacimiento"] or ""))
            self.table.setItem(row, 13, QTableWidgetItem(colaborator["numero_seguro_social"] or ""))
            self.table.setItem(row, 14, QTableWidgetItem(colaborator["informacion_adicional"]))

        self.result_label.setText(f"Mostrando {len(filtered_colaborators)} colaboradores filtrados")

    def export_to_excel_dialog(self):
        file_dialog = QFileDialog()
        file_path, _ = file_dialog.getSaveFileName(
            self,
            "Exportar a Excel",
            "",
            "Archivos Excel (*.xlsx)"
        )

        if file_path:
            if not file_path.lower().endswith(".xlsx"): # Verifica si termina en .xlsx (insensible a mayúsculas/minúsculas)
                file_path += ".xlsx" # Añade la extensión si falta

            try:
                self.export_table_to_excel(file_path)
                QMessageBox.information(self, "Exportar a Excel", "Tabla exportada exitosamente a Excel.")
            except Exception as e:
                QMessageBox.critical(self, "Error al exportar", f"Ocurrió un error al exportar a Excel: {str(e)}")

    def export_table_to_excel(self, file_path):
        """Exporta los datos del QTableWidget a un archivo Excel (.xlsx)."""
        libro_excel = openpyxl.Workbook()
        hoja_excel = libro_excel.active

        # Encabezados
        for columna in range(self.table.columnCount()):
            encabezado = self.table.horizontalHeaderItem(columna).text()
            hoja_excel.cell(row=1, column=columna + 1, value=encabezado)

        # Datos de la tabla
        for fila in range(self.table.rowCount()):
            for columna in range(self.table.columnCount()):
                item = self.table.item(fila, columna)
                if item and item.text():
                    dato = item.text()
                    hoja_excel.cell(row=fila + 2, column=columna + 1, value=dato)

        libro_excel.save(file_path)
        print(f"Tabla exportada a Excel: {file_path}")
