from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QTableWidget, QTableWidgetItem,
    QLineEdit, QPushButton, QLabel, QSizePolicy, QHeaderView,
    QFileDialog, QHBoxLayout, QMessageBox
)
from PyQt6.QtCore import Qt, QTimer
from PyQt6.QtGui import QIcon
import openpyxl

class WorkOrderTable(QWidget):
    def __init__(self, work_order_service, production_order_service):
        super().__init__()
        self.work_order_service = work_order_service
        self.production_order_service = production_order_service
        self.init_ui()

    def init_ui(self):
        self.setStyleSheet("background-color: white;")

        layout = QVBoxLayout()
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(5)

        filter_export_layout = QHBoxLayout()

        self.filter_input = QLineEdit(self)
        self.filter_input.setPlaceholderText("Filtrar por Numero de Orden o cliente...")
        self.filter_input.textChanged.connect(self.filter_table)
        filter_export_layout.addWidget(self.filter_input)

        self.export_excel_btn = QPushButton(self)
        self.export_excel_btn.setIcon(QIcon.fromTheme("document-save-as"))
        self.export_excel_btn.setText("Exportar a Excel")
        self.export_excel_btn.clicked.connect(self.export_to_excel_dialog)
        filter_export_layout.addWidget(self.export_excel_btn)

        layout.addLayout(filter_export_layout)

        self.table = QTableWidget(self)
        self.table.setColumnCount(11)
        self.table.setHorizontalHeaderLabels([
            "ID", "No Orden", "Fecha Inicio", "Fecha Fin", "ID Usuario", "ID Cliente",
            "ID Colaborador", "Costo Total", "Estado", "Servicios", "Comentario"
        ])
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)

        header = self.table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)

        layout.addWidget(self.table, stretch=1)

        self.refresh_btn = QPushButton("Refrescar Lista", self)
        self.refresh_btn.clicked.connect(self.load_work_orders)
        layout.addWidget(self.refresh_btn)

        self.result_label = QLabel(self)
        self.result_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(self.result_label)

        self.setLayout(layout)
        self.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)

        QTimer.singleShot(100, self.adjust_column_widths)
        self.load_work_orders()

    def adjust_column_widths(self):
        table_width = self.table.viewport().width()
        column_count = self.table.columnCount()
        if column_count > 0 and table_width > 0:
            equal_width = int(table_width / column_count)
            for col_index in range(column_count):
                self.table.setColumnWidth(col_index, equal_width)
        else:
            default_column_width = 20
            for col_index in range(column_count):
                self.table.setColumnWidth(col_index, default_column_width)

    def load_work_orders(self):
        work_orders = self.work_order_service.get_all_work_orders()
        self.table.setRowCount(0)

        if work_orders and isinstance(work_orders[0], tuple):
            work_order_keys = [
                "id", "work_order_id", "start_date", "end_date", "user_id", "client_id",
                "colaborator_id", "total_cost", "order_status", "note"
            ]
            work_orders = [dict(zip(work_order_keys, work_order)) for work_order in work_orders]

        for row, work_order in enumerate(work_orders):
            work_order_services_detail = self.work_order_service.get_work_order_items(str(work_order["work_order_id"]))
            services_str = work_order_services_detail[0][3]
            services_list = eval(services_str)
            services_text = '\n'.join(services_list)

            # Get Production order comment and handle None case
            production_order = self.production_order_service.get_production_order_details(str(work_order["work_order_id"]))
            note = production_order["note"] if production_order is not None else ""

            self.table.insertRow(row)
            self.table.setItem(row, 0, QTableWidgetItem(str(work_order["id"])))
            self.table.setItem(row, 1, QTableWidgetItem(str(work_order["work_order_id"])))
            self.table.setItem(row, 2, QTableWidgetItem(work_order["start_date"]))
            self.table.setItem(row, 3, QTableWidgetItem(work_order["end_date"]))
            self.table.setItem(row, 4, QTableWidgetItem(str(work_order["user_id"])))
            self.table.setItem(row, 5, QTableWidgetItem(str(work_order["client_id"])))
            self.table.setItem(row, 6, QTableWidgetItem(str(work_order["colaborator_id"])))
            self.table.setItem(row, 7, QTableWidgetItem(str(work_order["total_cost"])))
            self.table.setItem(row, 8, QTableWidgetItem(work_order["order_status"]))
            self.table.setItem(row, 9, QTableWidgetItem(services_text))
            self.table.setItem(row, 10, QTableWidgetItem(note))

        self.adjust_column_widths()
        self.result_label.setText(f"Mostrando {len(work_orders)} órdenes de trabajo")

    def filter_table(self):
        filter_text = self.filter_input.text().lower()
        all_work_orders = self.work_order_service.get_all_work_orders()

        if all_work_orders and isinstance(all_work_orders[0], tuple):
            work_order_keys = [
                "id", "work_order_id", "start_date", "end_date", "user_id", "client_id",
                "colaborator_id", "total_cost", "order_status", "note"
            ]
            all_work_orders = [dict(zip(work_order_keys, work_order)) for work_order in all_work_orders]

        filtered_work_orders = [
            work_order for work_order in all_work_orders
            if filter_text in str(work_order["work_order_id"]).lower() or filter_text in str(work_order["client_id"]).lower()
        ]

        self.table.setRowCount(0)

        for row, work_order in enumerate(filtered_work_orders):
            # Get and process services text
            work_order_services_detail = self.work_order_service.get_work_order_items(str(work_order["work_order_id"]))
            services_str = work_order_services_detail[0][3]
            services_list = eval(services_str)
            services_text = '\n'.join(services_list)

            # Get Production order comment and handle None case
            production_order = self.production_order_service.get_production_order_details(str(work_order["work_order_id"]))
            note = production_order["note"] if production_order is not None else ""

            self.table.insertRow(row)
            self.table.setItem(row, 0, QTableWidgetItem(str(work_order["id"])))
            self.table.setItem(row, 1, QTableWidgetItem(str(work_order["work_order_id"])))
            self.table.setItem(row, 2, QTableWidgetItem(work_order["start_date"]))
            self.table.setItem(row, 3, QTableWidgetItem(work_order["end_date"]))
            self.table.setItem(row, 4, QTableWidgetItem(str(work_order["user_id"])))
            self.table.setItem(row, 5, QTableWidgetItem(str(work_order["client_id"])))
            self.table.setItem(row, 6, QTableWidgetItem(str(work_order["colaborator_id"])))
            self.table.setItem(row, 7, QTableWidgetItem(str(work_order["total_cost"])))
            self.table.setItem(row, 8, QTableWidgetItem(work_order["order_status"]))
            self.table.setItem(row, 9, QTableWidgetItem(services_text))
            self.table.setItem(row, 10, QTableWidgetItem(note))

        self.result_label.setText(f"Mostrando {len(filtered_work_orders)} órdenes de trabajo filtradas")

    def export_to_excel_dialog(self):
        file_dialog = QFileDialog()
        file_path, _ = file_dialog.getSaveFileName(
            self,
            "Exportar a Excel",
            "",
            "Archivos Excel (*.xlsx)"
        )

        if file_path:
            if not file_path.lower().endswith(".xlsx"):
                file_path += ".xlsx"

            try:
                self.export_table_to_excel(file_path)
                QMessageBox.information(self, "Exportar a Excel", "Tabla exportada exitosamente a Excel.")
            except Exception as e:
                QMessageBox.critical(self, "Error al exportar", f"Ocurrió un error al exportar a Excel: {str(e)}")

    def export_table_to_excel(self, file_path):
        """Exporta los datos del QTableWidget a un archivo Excel (.xlsx)."""
        libro_excel = openpyxl.Workbook()
        hoja_excel = libro_excel.active

        # Encabezados
        for columna in range(self.table.columnCount()):
            encabezado = self.table.horizontalHeaderItem(columna).text()
            hoja_excel.cell(row=1, column=columna + 1, value=encabezado)

        # Datos de la tabla
        for fila in range(self.table.rowCount()):
            for columna in range(self.table.columnCount()):
                item = self.table.item(fila, columna)
                if item and item.text():
                    dato = item.text()
                    hoja_excel.cell(row=fila + 2, column=columna + 1, value=dato)

        libro_excel.save(file_path)
        print(f"Tabla exportada a Excel: {file_path}")
